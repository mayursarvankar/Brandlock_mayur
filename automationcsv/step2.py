import time

import pandas as pd
# import xlrd
import xlsxwriter
import datetime
import openpyxl
from win32com import client
#
import fitz
from step1 import *




file = 'Invoice Template test vdn4.xlsm'
validdf = firstStep()





def saveTopdf(InputFileName,saveFileName):
    xlApp = client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\automationcsv\\savedxlsx\\{InputFileName}.xlsx')
    ws = books.Worksheets[0]

    ws.Visible = 1
    ws.ExportAsFixedFormat(0, f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\automationcsv\\savedpdf\\{saveFileName}.pdf')

def removeExtrasheet(inputFileNmae):

    ipf = f"savedpdf\\{inputFileNmae}.pdf"
    opf = f"PDFs\\{inputFileNmae}.pdf"
    f = fitz.open(ipf)
    pgls = [0]
    f.select(pgls)
    f.save(opf)


#
invoiceNo=11050
for key in validdf:
    print("======================================")


    invoiceNo = invoiceNo+1
    siteNo=list(key.keys())[0]
    EppoSiteNo="Enoc / Eppco site " +siteNo

    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Invoice Template']

    ws['B10'] = EppoSiteNo
    ws['C3'] = str(invoiceNo)
    ws['C10'] = siteNo



    for values in key.values():
        DescriptioCellNo=18
        for inQty,itemName in (zip(values["InvoiceQty"] ,values['itemNames'])):
            DescriptioCellNo = DescriptioCellNo + 1

            ws[f"B{DescriptioCellNo}"]=itemName
            ws[f"F{DescriptioCellNo}"]=inQty


            print(itemName,inQty,DescriptioCellNo)

    wb.save(f"savedxlsx\\{invoiceNo}.xlsx")
    saveTopdf(invoiceNo, invoiceNo)
    removeExtrasheet(invoiceNo)
    wb.close()