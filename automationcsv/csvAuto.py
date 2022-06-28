import time

import pandas as pd
# import xlrd
import xlsxwriter

import openpyxl
from win32com import client
#
import fitz


file = 'Invoice Template test vdn3.xlsm'
df = pd.read_excel(file, sheet_name="Customers")
# ['Customer Name', 'Address 1', 'Address 2', 'Address 3', 'Address 4',
#        'Address 5', 'Email']

#

def saveTopdf(InputFileName,saveFileName):
    xlApp = client.Dispatch("Excel.Application")
    books = xlApp.Workbooks.Open(f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\csv\\savedxlsx\\{InputFileName}.xlsx')
    ws = books.Worksheets[0]

    ws.Visible = 1
    ws.ExportAsFixedFormat(0, f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\csv\\savedpdf\\{saveFileName}.pdf')

def removeExtrasheet(inputFileNmae):

    ipf = f"savedpdf\\{inputFileNmae}.pdf"
    opf = f"PDFs\\{inputFileNmae}.pdf"
    f = fitz.open(ipf)
    pgls = [0]
    f.select(pgls)
    f.save(opf)

#
invoiceNo=2000
for custName in df['Customer Name'].values:
    invoiceNo = invoiceNo +1

    print(invoiceNo,custName)

    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Invoice Template']

    ws['B10'] = custName
    ws['C3']=str(invoiceNo)
    wb.save(f"savedxlsx\\{invoiceNo}.xlsx")

    saveTopdf(invoiceNo, invoiceNo)
    removeExtrasheet(invoiceNo)
    wb.close()


    # if invoiceNo ==2002:break



#
#

# xlApp = client.Dispatch("Excel.Application")
# books = xlApp.Workbooks.Open(f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\csv\\savedxlsx\\Invoice Template test vdn3.xlsm')
# ws = books.Worksheets[0]
#
# ws.Visible = 1
# ws.ExportAsFixedFormat(0, f'C:\\Users\\Admin\\PycharmProjects\\db_conection_screenshot\\csv\\savedpdf\\demo.pdf')

